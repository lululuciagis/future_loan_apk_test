
import openpyxl
import requests
import jsonpath
#定义一个python读取excel列表的函数
def read_excel(workbookname,sheetname):
    # 工程中放入excel表格
    #加载工作簿
    workbook = openpyxl.load_workbook(workbookname)
    #获取表单
    sheet = workbook[sheetname]
    data_api_test = [] #创建一个空列表，用来存放所有测试用例
    #获取单元格
    for i in range(2,sheet.max_row+1):#注意range()，读头不读尾，这里取值max_row+1
        dict_excel = \
            dict(
            case_id_excel = sheet.cell(row=i,column=1).value,    #获取case_id
            url_excel = sheet.cell(row=i,column=5).value,    #获取url
            data_excel = sheet.cell(row = i,column = 6).value,  #获取data
            expected_excel = sheet.cell(row = i,column = 7).value   #获取expected
            )
        data_api_test.append(dict_excel)  #循环一次，追加一次生成的字典到列表中
    #print(data_api_test) #打印最终列表结果
    return data_api_test
#定义一个接口测试的函数，post请求方式
def api_test(url1,json1):
    result = requests.post(url=url1, json=json1, headers={'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'})
    return result.json()
#定义一个python写数据至excel中的函数
def write_excel(workbookname,sheetname,row1,column1,content):
    #加载工作簿
    workbook = openpyxl.load_workbook(workbookname)
    #获取表单
    sheet = workbook[sheetname]
    #写入结果
    sheet.cell(row=row1,column=column1).value=content
    workbook.save(workbookname)
cases = read_excel('api_test.xlsx','register')  #调用函数读取excel中的测试用例
#print(cases)
#print(cases)
for case in cases:
    id_case = case['case_id_excel']   #获取每一条测试用例的id
    print(id_case)
    url_case = case.get('url_excel')   #获取每一条测试用例的url
    json1_case = eval(case.get('data_excel'))   #获取每一条测试用例的请求体,通过eval()将字符串获取为字典格式
    print(type(json1_case))
    expected_case = eval(case.get('expected_excel'))  #获取每一条测试用例的预期值
    api_test_result = api_test(url1= url_case,json1=json1_case)   #调用函数运行接口测试
    api_test_result_interest = api_test_result.get('msg')  #从接口测试结果中获取msg值
    if api_test_result_interest == expected_case.get('msg'):      #断言
        print('第{}条测试用例通过'.format(id_case))
        final_result = 'pass'
    else:
        print('第{}条测试用例没有通过'.format(id_case))
        final_result= 'failed'
    write_excel('api_test.xlsx', 'register', id_case + 1, 8, final_result)  # 调用write_excel()函数实现数据写入
    print('——'*30)
